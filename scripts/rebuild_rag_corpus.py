#!/usr/bin/env python3
"""Rebuild the metacoupling RAG markdown corpus.

This script implements the mixed copyright strategy:

- ``ship=yes`` rows are copied as stripped full-text markdown.
- ``ship=no`` rows are converted into short, paraphrased records generated
  by two OpenAI calls: summary, then structured metacoupling extraction.

The script is resumable. LLM outputs are cached in ``tmp/rag_rebuild_cache``
by default, so reruns do not repeat API calls unless ``--force`` is used.

Example dry run:

    python scripts/rebuild_rag_corpus.py --dry-run

Validate one closed paper:

    python scripts/rebuild_rag_corpus.py --only closed --limit 1 \
        --env-file "D:/Onedrive/OneDrive - Michigan State University/Desktop/api.env"

Full rebuild:

    python scripts/rebuild_rag_corpus.py \
        --env-file "D:/Onedrive/OneDrive - Michigan State University/Desktop/api.env"
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_TABLE = ROOT / "tmp" / "strategy_b_decision.xlsx"
DEFAULT_SOURCE_ROOT = Path(r"D:\Allpapers\_markdown_stripped\Journal Articles")
DEFAULT_OUTPUT_DIR = ROOT / "Papers_rebuilt"
DEFAULT_CACHE_DIR = ROOT / "tmp" / "rag_rebuild_cache"
DEFAULT_MODEL = "gpt-5.1"
PROMPT_VERSION = "rag_rebuild_v4"


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


@dataclass
class PaperRow:
    ship: str
    reason: str
    filename: str
    title: str
    year: str
    doi: str
    oa_status: str
    license: str
    category: str


def _parse_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            values[key] = value
    return values


def _load_api_key(env_file: Path | None) -> str:
    if env_file is not None:
        values = _parse_env_file(env_file)
        if values.get("OPENAI_API_KEY"):
            return values["OPENAI_API_KEY"]
    key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not key:
        raise RuntimeError(
            "OPENAI_API_KEY was not found. Pass --env-file or set the "
            "environment variable."
        )
    return key


def _load_rows(table: Path) -> list[PaperRow]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(table, read_only=True, data_only=True)
    ws = wb["Journal Articles"]
    raw_headers = next(ws.iter_rows(values_only=True))
    headers = [str(x).strip() if x is not None else "" for x in raw_headers]
    idx = {h: i for i, h in enumerate(headers)}
    required = {
        "ship", "reason", "filename", "title", "year",
        "doi", "oa_status", "license", "category",
    }
    missing = sorted(required - set(idx))
    if missing:
        raise RuntimeError(f"Missing required columns in {table}: {missing}")

    rows: list[PaperRow] = []
    for values in ws.iter_rows(values_only=True):
        filename = str(values[idx["filename"]] or "").strip()
        if not filename or filename.lower() == "filename":
            continue
        rows.append(PaperRow(
            ship=str(values[idx["ship"]] or "").strip().lower(),
            reason=str(values[idx["reason"]] or "").strip(),
            filename=filename,
            title=str(values[idx["title"]] or "").strip(),
            year=str(values[idx["year"]] or "").strip(),
            doi=str(values[idx["doi"]] or "").strip(),
            oa_status=str(values[idx["oa_status"]] or "").strip(),
            license=str(values[idx["license"]] or "").strip(),
            category=str(values[idx["category"]] or "").strip(),
        ))
    return rows


def _build_markdown_index(source_root: Path) -> dict[str, Path]:
    index: dict[str, Path] = {}
    for path in source_root.rglob("*.md"):
        index[path.name.lower()] = path
        index[path.stem.lower()] = path
        index[_filename_match_key(path.name)] = path
        index[_filename_match_key(path.stem)] = path
    return index


def _filename_match_key(value: str) -> str:
    """Return a permissive ASCII key for mojibake/non-ASCII filename matching."""
    value = value.lower()
    value = value.encode("ascii", errors="ignore").decode("ascii")
    value = re.sub(r"\.(pdf|md|txt)$", "", value)
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def _find_markdown(row: PaperRow, index: dict[str, Path]) -> Path | None:
    pdf_path = Path(row.filename)
    candidates = [
        row.filename,
        pdf_path.name,
        pdf_path.stem,
        pdf_path.with_suffix(".md").name,
        f"{pdf_path.stem}.md",
    ]
    for candidate in candidates:
        hit = index.get(candidate.lower()) or index.get(_filename_match_key(candidate))
        if hit is not None:
            return hit
    return None


def _infer_title(row: PaperRow, md_path: Path | None = None) -> str:
    if row.title:
        return row.title
    stem = Path(row.filename).stem
    match = re.match(r"^.+?\s+-\s+\d{4}\s+-\s+(.+)$", stem)
    if match:
        return match.group(1).strip()
    if md_path is not None:
        return md_path.stem
    return stem


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()


def _cache_path(
    cache_dir: Path,
    row: PaperRow,
    kind: str,
    model: str,
    source_text: str,
) -> Path:
    payload = json.dumps(
        {
            "prompt_version": PROMPT_VERSION,
            "kind": kind,
            "model": model,
            "filename": row.filename,
            "doi": row.doi,
            "source_hash": _sha256_text(source_text),
        },
        sort_keys=True,
    )
    digest = _sha256_text(payload)[:20]
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(row.filename).stem)[:90]
    suffix = "json" if kind == "extraction" else "md"
    return cache_dir / f"{safe_stem}.{kind}.{digest}.{suffix}"


def _extract_output_text(response: Any) -> str:
    output_text = getattr(response, "output_text", None)
    if isinstance(output_text, str):
        return output_text.strip()

    pieces: list[str] = []
    for item in getattr(response, "output", []) or []:
        for content in getattr(item, "content", []) or []:
            text = getattr(content, "text", None)
            if text:
                pieces.append(str(text))
    return "\n".join(pieces).strip()


def _call_openai(
    client: Any,
    *,
    model: str,
    system: str,
    user: str,
    max_output_tokens: int,
    retries: int = 4,
) -> str:
    last_exc: Exception | None = None
    for attempt in range(retries):
        try:
            if hasattr(client, "responses"):
                response = client.responses.create(
                    model=model,
                    input=[
                        {"role": "system", "content": system},
                        {"role": "user", "content": user},
                    ],
                    max_output_tokens=max_output_tokens,
                )
                text = _extract_output_text(response)
            else:
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": system},
                        {"role": "user", "content": user},
                    ],
                    max_completion_tokens=max_output_tokens,
                )
                text = response.choices[0].message.content or ""
            if text.strip():
                return text.strip()
            raise RuntimeError("OpenAI returned empty text.")
        except Exception as exc:
            last_exc = exc
            wait = min(60, 2 ** attempt)
            time.sleep(wait)
    raise RuntimeError(f"OpenAI call failed after {retries} attempts: {last_exc}")


def _call_structured_extraction(
    client: Any,
    *,
    model: str,
    system: str,
    user: str,
    max_output_tokens: int,
) -> str:
    """Call OpenAI for structured extraction, retrying once with more room."""
    try:
        text = _call_openai(
            client,
            model=model,
            system=system,
            user=user,
            max_output_tokens=max_output_tokens,
        )
        extraction = _load_jsonish(text)
        _validate_extraction(extraction)
        return text
    except ValueError:
        retry_user = (
            user
            + "\n\nRetry instruction: The previous answer was invalid or "
            "too long. Return compact valid JSON only. Use at most 2 strings "
            "per list and keep each string under 12 words."
        )
        retry_tokens = max(max_output_tokens, 6000)
        text = _call_openai(
            client,
            model=model,
            system=system,
            user=retry_user,
            max_output_tokens=retry_tokens,
        )
        extraction = _load_jsonish(text)
        _validate_extraction(extraction)
        return text


def _paper_metadata_block(row: PaperRow, title: str) -> str:
    return "\n".join([
        f"Title: {title}",
        f"Filename: {row.filename}",
        f"Year: {row.year or 'unknown'}",
        f"DOI: {row.doi or 'unknown'}",
        f"OA status: {row.oa_status or 'unknown'}",
        f"License: {row.license or 'none recorded'}",
        f"Category: {row.category or 'unknown'}",
        f"Closed-access reason: {row.reason or 'not recorded'}",
    ])


def _summary_prompt(row: PaperRow, title: str, source_text: str) -> tuple[str, str]:
    system = (
        "You create copyright-safe research summaries for a local RAG corpus. "
        "Do not copy sentences from the source. Do not quote. Do not reproduce "
        "tables, abstracts, figures, or long method/result passages. Write in "
        "English even when the source is Chinese or Spanish. Use plain ASCII "
        "punctuation only. The output must be a paraphrased summary record, "
        "not a substitute for the paper."
    )
    user = f"""\
Create a concise English summary for this paper.

Rules:
- 250 to 350 words.
- Paraphrase everything.
- No direct quotes.
- Use plain ASCII punctuation only: straight apostrophes, straight quotes,
  and hyphens instead of curly punctuation.
- Include study topic, geographic scope, system context, methods/data if clear,
  main findings, and relevance to telecoupling/metacouplingllm.
- If the paper is not actually about metacoupling/telecoupling, say so briefly.

Metadata:
{_paper_metadata_block(row, title)}

Source text:
<<<SOURCE_TEXT
{source_text}
SOURCE_TEXT
>>>
"""
    return system, user


def _extraction_prompt(row: PaperRow, title: str, source_text: str) -> tuple[str, str]:
    system = (
        "You extract structured metacoupling information for a RAG corpus. "
        "Use only paraphrased content. Do not quote from the paper. Return "
        "only compact valid JSON. Write all values in English. Use plain "
        "ASCII punctuation only."
    )
    schema = {
        "language_detected": "English, Chinese, Spanish, etc.",
        "coupling_classification": [
            "Brief classification statement for intra/peri/telecoupling relevance."
        ],
        "intracoupling": {
            "systems": [],
            "flows": [],
            "agents": [],
            "causes": [],
            "effects": [],
        },
        "pericoupling": {
            "systems": [],
            "flows": [],
            "agents": [],
            "causes": [],
            "effects": [],
        },
        "telecoupling": {
            "systems": [],
            "flows": [],
            "agents": [],
            "causes": [],
            "effects": [],
        },
        "cross_coupling_interactions": [],
        "research_gaps_or_limitations": [],
        "search_keywords": [],
        "confidence": "high, medium, or low",
    }
    user = f"""\
Extract metacoupling-relevant information from this paper.

Rules:
- Return only JSON matching this schema.
- Use short paraphrased phrases or bullet-like strings.
- Keep every list to at most 3 strings.
- Keep each string under 18 words.
- Use empty lists when a component is not supported by the source.
- Do not quote the source text.
- Include Spanish/Chinese content as English paraphrase.
- Use plain ASCII punctuation only.
- Keep the whole JSON concise: target 250 to 450 words total.

Schema:
{json.dumps(schema, indent=2)}

Metadata:
{_paper_metadata_block(row, title)}

Source text:
<<<SOURCE_TEXT
{source_text}
SOURCE_TEXT
>>>
"""
    return system, user


def _load_jsonish(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        value = json.loads(cleaned)
        return value if isinstance(value, dict) else {"raw": value}
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
        if match:
            try:
                value = json.loads(match.group(0))
                return value if isinstance(value, dict) else {"raw": value}
            except json.JSONDecodeError:
                pass
    preview = cleaned[:300].replace("\n", " ")
    raise ValueError(f"Structured extraction was not valid JSON: {preview}")


def _validate_extraction(extraction: dict[str, Any]) -> None:
    """Fail fast on empty/low-quality structured extraction records."""
    content_fields: list[Any] = [
        extraction.get("coupling_classification"),
        extraction.get("cross_coupling_interactions"),
        extraction.get("research_gaps_or_limitations"),
        extraction.get("search_keywords"),
    ]
    for block_name in ("intracoupling", "pericoupling", "telecoupling"):
        block = extraction.get(block_name)
        if isinstance(block, dict):
            content_fields.extend(
                block.get(field)
                for field in ("systems", "flows", "agents", "causes", "effects")
            )

    def has_content(value: Any) -> bool:
        if isinstance(value, str):
            return bool(value.strip())
        if isinstance(value, list):
            return any(has_content(item) for item in value)
        if isinstance(value, dict):
            return any(has_content(item) for item in value.values())
        return value is not None

    if not any(has_content(value) for value in content_fields):
        raise ValueError(
            "Structured extraction was valid JSON but contained no usable "
            "metacoupling content."
        )


def _format_list(values: Any) -> str:
    if not values:
        return "- Not identified."
    if isinstance(values, str):
        return f"- {values.strip()}"
    if isinstance(values, list):
        lines = []
        for item in values:
            if isinstance(item, str):
                lines.append(f"- {item.strip()}")
            else:
                lines.append(f"- {json.dumps(item, ensure_ascii=False)}")
        return "\n".join(lines) if lines else "- Not identified."
    return f"- {json.dumps(values, ensure_ascii=False)}"


def _format_component_block(name: str, value: Any) -> str:
    if not isinstance(value, dict):
        return f"### {name}\n{_format_list(value)}\n"
    lines = [f"### {name}"]
    for field in ("systems", "flows", "agents", "causes", "effects"):
        lines.append(f"#### {field.title()}")
        lines.append(_format_list(value.get(field, [])))
    return "\n".join(lines) + "\n"


def _write_closed_record(
    output_path: Path,
    row: PaperRow,
    title: str,
    summary: str,
    extraction: dict[str, Any],
) -> None:
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    lines = [
        f"# {title}",
        "",
        "<!--",
        "RAG record type: paraphrased-summary-and-structured-extraction",
        "Copyright note: This file intentionally contains no full text from",
        "the source paper. It is a short paraphrased metadata/summary record",
        "for retrieval only.",
        f"Generated at: {now}",
        f"Prompt version: {PROMPT_VERSION}",
        "-->",
        "",
        "## Metadata",
        f"- Filename: {row.filename}",
        f"- Year: {row.year or 'unknown'}",
        f"- DOI: {row.doi or 'unknown'}",
        f"- OA status: {row.oa_status or 'unknown'}",
        f"- License: {row.license or 'none recorded'}",
        f"- Category: {row.category or 'unknown'}",
        f"- Closed-access reason: {row.reason or 'not recorded'}",
        "",
        "## Copyright-Safe Summary",
        summary.strip(),
        "",
        "## Structured Metacoupling Extraction",
        "",
        "### Coupling Classification",
        _format_list(extraction.get("coupling_classification", [])),
        "",
        _format_component_block("Intracoupling", extraction.get("intracoupling", {})),
        _format_component_block("Pericoupling", extraction.get("pericoupling", {})),
        _format_component_block("Telecoupling", extraction.get("telecoupling", {})),
        "### Cross-Coupling Interactions",
        _format_list(extraction.get("cross_coupling_interactions", [])),
        "",
        "### Research Gaps or Limitations",
        _format_list(extraction.get("research_gaps_or_limitations", [])),
        "",
        "### Search Keywords",
        _format_list(extraction.get("search_keywords", [])),
        "",
        "### Extraction Confidence",
        str(extraction.get("confidence", "not reported")),
        "",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def _copy_open_record(source_path: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)


def rebuild(args: argparse.Namespace) -> int:
    rows = _load_rows(args.table)
    index = _build_markdown_index(args.source_root)
    output_dir: Path = args.output_dir
    cache_dir: Path = args.cache_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)

    if args.only == "open":
        rows = [r for r in rows if r.ship == "yes"]
    elif args.only == "closed":
        rows = [r for r in rows if r.ship != "yes"]
    if args.start_at is not None:
        if args.start_at < 1:
            raise RuntimeError("--start-at must be 1 or greater.")
        rows = rows[args.start_at - 1:]
    if args.limit is not None:
        rows = rows[: args.limit]

    client = None
    if not args.dry_run and any(r.ship != "yes" for r in rows):
        api_key = _load_api_key(args.env_file)
        try:
            from openai import OpenAI
        except ImportError as exc:
            raise RuntimeError("openai is required: pip install openai") from exc
        client = OpenAI(api_key=api_key)

    manifest: dict[str, Any] = {
        "started_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_root": str(args.source_root),
        "table": str(args.table),
        "output_dir": str(output_dir),
        "cache_dir": str(cache_dir),
        "model": args.model,
        "prompt_version": PROMPT_VERSION,
        "dry_run": bool(args.dry_run),
        "counts": {
            "selected_rows": len(rows),
            "open_copied": 0,
            "closed_generated": 0,
            "skipped_existing": 0,
            "missing_source": 0,
            "errors": 0,
        },
        "missing": [],
        "errors": [],
    }

    for n, row in enumerate(rows, 1):
        source_path = _find_markdown(row, index)
        if source_path is None:
            manifest["counts"]["missing_source"] += 1
            manifest["missing"].append(row.filename)
            print(f"[{n}/{len(rows)}] MISSING: {row.filename}")
            continue

        output_path = output_dir / source_path.name
        title = _infer_title(row, source_path)

        if output_path.exists() and not args.force:
            manifest["counts"]["skipped_existing"] += 1
            print(f"[{n}/{len(rows)}] SKIP existing: {output_path.name}")
            continue

        if row.ship == "yes":
            print(f"[{n}/{len(rows)}] COPY open full text: {source_path.name}")
            if not args.dry_run:
                _copy_open_record(source_path, output_path)
            manifest["counts"]["open_copied"] += 1
            continue

        print(f"[{n}/{len(rows)}] SUMMARIZE closed paper: {source_path.name}")
        if args.dry_run:
            manifest["counts"]["closed_generated"] += 1
            continue

        try:
            source_text = source_path.read_text(encoding="utf-8", errors="ignore")

            summary_cache = _cache_path(
                cache_dir, row, "summary", args.model, source_text,
            )
            if summary_cache.exists() and not args.force:
                summary = summary_cache.read_text(encoding="utf-8")
            else:
                system, user = _summary_prompt(row, title, source_text)
                summary = _call_openai(
                    client,
                    model=args.model,
                    system=system,
                    user=user,
                    max_output_tokens=args.summary_tokens,
                )
                summary_cache.write_text(summary, encoding="utf-8")
                time.sleep(args.sleep)

            extraction_cache = _cache_path(
                cache_dir, row, "extraction", args.model, source_text,
            )
            if extraction_cache.exists() and not args.force:
                extraction_text = extraction_cache.read_text(encoding="utf-8")
                extraction = _load_jsonish(extraction_text)
                _validate_extraction(extraction)
            else:
                system, user = _extraction_prompt(row, title, source_text)
                extraction_text = _call_structured_extraction(
                    client,
                    model=args.model,
                    system=system,
                    user=user,
                    max_output_tokens=args.extraction_tokens,
                )
                extraction_cache.write_text(extraction_text, encoding="utf-8")
                time.sleep(args.sleep)

            extraction = _load_jsonish(extraction_text)
            _validate_extraction(extraction)
            _write_closed_record(output_path, row, title, summary, extraction)
            manifest["counts"]["closed_generated"] += 1
        except Exception as exc:
            manifest["counts"]["errors"] += 1
            err = {"filename": row.filename, "error": str(exc)}
            manifest["errors"].append(err)
            print(f"ERROR: {row.filename}: {exc}", file=sys.stderr)
            if args.stop_on_error:
                break

    manifest["finished_at"] = dt.datetime.now(dt.timezone.utc).isoformat()
    manifest_path = output_dir / "rag_rebuild_manifest.json"
    if not args.dry_run:
        manifest_path.write_text(
            json.dumps(manifest, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    else:
        print(json.dumps(manifest["counts"], indent=2))
    print("Done.")
    print(json.dumps(manifest["counts"], indent=2))
    return 0 if manifest["counts"]["errors"] == 0 else 1


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Rebuild the metacoupling RAG corpus.",
    )
    parser.add_argument("--table", type=Path, default=DEFAULT_TABLE)
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE_DIR)
    parser.add_argument("--env-file", type=Path, default=None)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument(
        "--only",
        choices=("all", "open", "closed"),
        default="all",
        help="Process only open/full-text rows, closed/summary rows, or all.",
    )
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument(
        "--start-at",
        type=int,
        default=None,
        help="Start processing at this 1-based row number after --only filtering.",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--stop-on-error", action="store_true")
    parser.add_argument("--sleep", type=float, default=0.3)
    parser.add_argument("--summary-tokens", type=int, default=2500)
    parser.add_argument("--extraction-tokens", type=int, default=4000)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    return rebuild(args)


if __name__ == "__main__":
    raise SystemExit(main())
