"""Cross-check the 'Journal Articles/Research' markdown folder against
the .bib file.

For each .md paper, verify:
  1. It has a .bib entry (by bib_key mapping from CSV + manual overrides).
  2. The .bib entry carries a citation count (annote = {Cited by: N}).
  3. If the paper has user-provided keywords (from either
     journal_research Excel or Check.xlsx), the .bib keywords match.

Reports all mismatches to tmp/verify_papers_vs_bib.txt.
"""
from __future__ import annotations

import csv
import glob
import re
from pathlib import Path

import openpyxl

RESEARCH_DIR = Path(r"D:/Check/_markdown/Journal Articles/Research")
BIB_PATH = Path("src/metacoupling/data/telecoupling_literature.bib")
CSV_PATH = Path("tmp/keyword_extraction.csv")
JOURNAL_XLSX = Path("D:/Check/journal_research_keywords_extracted.xlsx")
CHECK_XLSX = Path(
    r"D:/Onedrive/OneDrive - Michigan State University/Desktop/Check.xlsx"
)
OUT = Path("tmp/verify_papers_vs_bib.txt")

# Manual overrides learned across the investigation (filename -> bib_key).
# These are the 11 files whose Zotero-generated filename didn't match the
# .bib key convention.
MANUAL_OVERRIDES = {
    "Javier Alcantara-Plazola and de la Barrera - 2021 - Quantification of embedded phosphorus in Mexican agriculture.md": "alcntaraplazola_quantification_2021",
    "Bicudo da Silva et al. - 2019 - Eco-certification protocols as mechanisms to foster sustainable environmental practices in telecoupl.md": "dasilva_eco_2019",
    "Yue et al. - 2018 - Spillover effect offsets the conservation effort in the Amazon.md": "dou_spillover_2018",
    "Du et al. - 2022 - How far are we from possible ideal virtual water transfer Evidence from assessing vulnerability of.md": "du_how_2022_a",
    "ChuangLin and Yufei - 2017 - Analysis of emergy-based metabolic efficiency and environmental pressure on the local coupling and t.md": "fang_analysis_2017",
    "Friis and Nielsen - 2017 - On the System. Boundary Choices, Implications, and Solutions in Telecoupling Land Use Change Researc.md": "friis_system_2017",
    "Lopez-Hoffman et al. - 2017 - Operationalizing the telecoupling framework for migratory species using the spatial subsidies approa.md": "lpezhoffman_operationalizing_2017",
    "de Lucio et al. - 2021 - Resilience as a Moving Target An Evaluation of Last Century Management Strategies in a Dry-Edge Mar.md": "morenofernndez_resilience_2021",
    "Rey and Huettmann - 2020 - Telecoupling analysis of the Patagonian Shelf A new approach to study global seabird-fisheries inte.md": "rayarey_telecoupling_2020",
    "Sondergaard et al. - 2024 - Fragmented sustainability governance of telecoupled flows Brazilian beef exports to China.md": "sndergaard_fragmented_2024",
    "Vila and Arzamendia - 2022 - South American Camelids their values and contributions to people.md": "vil_south_2022",
    # Drakou-filename is the marston paper (Zotero metadata quirk)
    "Drakou et al. - 2017 - Drought impacts to water footprints and virtual water transfers of the Central Valley of California.md": "marston_drought_2017",
    # Already-known overrides (kept for completeness)
    "Liu and Pan - 2025 - Unraveling the quantity and sustainability of cross-scale ecosystem service flows A meta-coupling f.md": "liu_unraveling_2025_a",
    "Liu et al. - 2025 - From Plate to Plow How Dietary Shifts Drive Telecoupled Cropland Erosion in China.md": "liu_plate_2025",
    "Li et al. - 2019 - Tightening ecological management facilitates green development in the Qilian Mountains.md": "zhao_tightening_2019",
    "Reis et al. - 2020 - Understanding the Stickiness of Commodity Supply Chains Is Key to Improving Their Sustainability.md": "dosreis_understanding_2020",
    # Zheng and Sun uses underscore separator, not dash
    "Zheng and Sun_The short-range and remote analysis of virtual water trade in China.md": "zheng_short_2023",
}


def extract_field_balanced(chunk: str, field: str) -> str | None:
    """Extract a bibtex field value handling nested/escaped braces (e.g. LaTeX)."""
    m = re.search(rf"\b{field}\s*=\s*\{{", chunk, re.IGNORECASE)
    if not m:
        return None
    i = m.end()
    depth = 1
    while i < len(chunk) and depth > 0:
        c = chunk[i]
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return chunk[m.end():i]
        i += 1
    return None


def parse_bib(text: str) -> dict[str, dict]:
    starts = list(re.finditer(r"^@\w+\{([^,]+),", text, re.MULTILINE))
    out = {}
    for i, m in enumerate(starts):
        end = starts[i + 1].start() if i + 1 < len(starts) else len(text)
        chunk = text[m.start():end]
        key = m.group(1).strip()
        annote = extract_field_balanced(chunk, "annote") or ""
        cite_m = re.search(r"Cited\s*by:\s*(\d+)", annote, re.IGNORECASE)
        kw = extract_field_balanced(chunk, "keywords")
        title = extract_field_balanced(chunk, "title") or ""
        out[key] = {
            "citation": int(cite_m.group(1)) if cite_m else None,
            "has_citation": cite_m is not None,
            "has_keywords": kw is not None,
            "keywords": (kw or "").strip(),
            "title": title.strip(),
        }
    return out


def load_csv_map() -> dict[str, str]:
    out = {}
    with open(CSV_PATH, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r.get("bib_key"):
                out[r["filename"]] = r["bib_key"]
    return out


def load_journal_xlsx() -> dict[str, dict]:
    """filename -> {keywords, status} from journal_research_keywords_extracted."""
    wb = openpyxl.load_workbook(JOURNAL_XLSX, read_only=True, data_only=True)
    sh = wb["All Journal Research"]
    header = [c.value for c in next(sh.iter_rows())]
    out = {}
    for row in sh.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        fn = d.get("filename") or ""
        if fn:
            out[fn] = {
                "keywords": (d.get("keywords") or "").strip(),
                "status": (d.get("keyword_status") or "").strip(),
            }
    return out


def load_check_xlsx() -> dict[str, str]:
    """bib_key -> keywords (or 'NA') from Check.xlsx."""
    wb = openpyxl.load_workbook(CHECK_XLSX, read_only=True, data_only=True)
    sh = wb["Sheet1"]
    out = {}
    for i, r in enumerate(sh.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not r or not r[0]:
            continue
        out[str(r[0]).strip()] = str(r[1]).strip() if r[1] else ""
    return out


def tokenize_kw(s: str) -> set[str]:
    """Lowercase, punctuation-strip set of items for comparison."""
    items = re.split(r"[;,]", s)
    out = set()
    for it in items:
        t = re.sub(r"[^a-z0-9 ]", " ", it.lower())
        t = re.sub(r"\s+", " ", t).strip()
        if t:
            out.add(t)
    return out


def main():
    md_files = sorted(glob.glob(str(RESEARCH_DIR / "**/*.md"), recursive=True))
    print(f"Markdown files in Research folder: {len(md_files)}")

    bib = parse_bib(BIB_PATH.read_text(encoding="utf-8"))
    print(f".bib entries: {len(bib)}")

    csv_map = load_csv_map()
    csv_map.update(MANUAL_OVERRIDES)

    journal_xlsx = load_journal_xlsx()
    check_xlsx = load_check_xlsx()

    # Issues
    no_bib: list[tuple[str, str]] = []       # (filename, reason)
    no_citation: list[tuple[str, str]] = []  # (bib_key, filename)
    kw_mismatch: list[dict] = []             # detailed info
    kw_missing_in_bib: list[dict] = []       # user provided, .bib empty
    kw_extra_in_bib: list[dict] = []         # .bib has kw, user says none

    for md in md_files:
        fn = Path(md).name
        key = csv_map.get(fn)
        if not key:
            no_bib.append((fn, "no CSV mapping"))
            continue
        if key not in bib:
            no_bib.append((fn, f"CSV maps to '{key}' but key absent from .bib"))
            continue

        b = bib[key]
        if not b["has_citation"]:
            no_citation.append((key, fn))

        # Determine what kw user has for this paper.
        user_kw = ""
        user_has = False  # True if user explicitly provided kw
        user_explicitly_none = False  # True if user said NA/no-kw
        if key in check_xlsx:
            v = check_xlsx[key].upper()
            if v == "NA" or not check_xlsx[key]:
                user_explicitly_none = True
            else:
                user_kw = check_xlsx[key]
                user_has = True
        elif fn in journal_xlsx:
            jx = journal_xlsx[fn]
            if jx["keywords"]:
                user_kw = jx["keywords"]
                user_has = True
            else:
                user_explicitly_none = True

        if user_has:
            if not b["keywords"]:
                kw_missing_in_bib.append({
                    "key": key, "fn": fn, "user_kw": user_kw,
                })
            else:
                ut = tokenize_kw(user_kw)
                bt = tokenize_kw(b["keywords"])
                # Consider a match if 80%+ of user tokens are in bib
                if not ut:
                    continue
                inter = len(ut & bt)
                if inter / len(ut) < 0.8:
                    kw_mismatch.append({
                        "key": key, "fn": fn,
                        "user_kw": user_kw,
                        "bib_kw": b["keywords"],
                        "coverage": round(inter / len(ut), 2),
                    })
        elif user_explicitly_none and b["has_keywords"]:
            kw_extra_in_bib.append({
                "key": key, "fn": fn,
                "bib_kw": b["keywords"],
            })

    # Also check for .bib entries with no corresponding markdown
    mapped_keys = {csv_map[fn] for fn in (Path(m).name for m in md_files) if fn in csv_map}
    bib_keys_without_md = sorted(set(bib) - mapped_keys)

    lines: list[str] = []
    pr = lines.append

    pr(f"Markdown files: {len(md_files)}")
    pr(f".bib entries:   {len(bib)}")
    pr(f"Mapped keys:    {len(mapped_keys)}")
    pr(f"")
    pr(f"SUMMARY")
    pr(f"  Markdown files without .bib entry:       {len(no_bib)}")
    pr(f"  .bib entries missing citation count:     {len(no_citation)}")
    pr(f"  Keyword mismatches (coverage < 80%):     {len(kw_mismatch)}")
    pr(f"  User has kw but .bib has none:           {len(kw_missing_in_bib)}")
    pr(f"  .bib has kw but user says no kw (NA):    {len(kw_extra_in_bib)}")
    pr(f"  .bib keys without a markdown in folder:  {len(bib_keys_without_md)}")
    pr(f"")

    def section(title: str):
        pr("=" * 78)
        pr(title)
        pr("=" * 78)

    if no_bib:
        section(f"1. Markdown files with no .bib mapping  ({len(no_bib)})")
        for fn, reason in no_bib:
            pr(f"  - {fn}")
            pr(f"      reason: {reason}")
        pr("")

    if no_citation:
        section(f"2. .bib entries missing citation count  ({len(no_citation)})")
        for key, fn in no_citation:
            pr(f"  - {key}")
            pr(f"      markdown: {fn[:110]}")
        pr("")

    if kw_missing_in_bib:
        section(f"3. User has kw but .bib has none  ({len(kw_missing_in_bib)})")
        for d in kw_missing_in_bib:
            pr(f"  - {d['key']}")
            pr(f"      user: {d['user_kw'][:110]}")
        pr("")

    if kw_extra_in_bib:
        section(f"4. .bib has kw but user marked NA  ({len(kw_extra_in_bib)})")
        for d in kw_extra_in_bib:
            pr(f"  - {d['key']}")
            pr(f"      bib: {d['bib_kw'][:110]}")
        pr("")

    if kw_mismatch:
        section(f"5. Keyword content mismatch (coverage < 80%)  ({len(kw_mismatch)})")
        for d in kw_mismatch:
            pr(f"  - {d['key']}  (coverage {d['coverage']})")
            pr(f"      user: {d['user_kw'][:110]}")
            pr(f"      bib:  {d['bib_kw'][:110]}")
        pr("")

    if bib_keys_without_md:
        section(f"6. .bib keys without matching markdown  ({len(bib_keys_without_md)})")
        for k in bib_keys_without_md:
            b = bib[k]
            pr(f"  - {k}")
            pr(f"      title: {b['title'][:110]}")
            pr(f"      citation: {b['citation']}  has_kw: {b['has_keywords']}")
        pr("")

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT}")
    # Also print summary to stdout
    print()
    for line in lines[:10]:
        print(line)


if __name__ == "__main__":
    main()
