"""Audit the 16 DELETED + 20 PRESERVED bib entries for recoverable
Excel keywords that failed to match during integration.

For each entry, try to find a user-provided keyword row in the Excel
whose markdown filename plausibly corresponds. Flags entries where the
user HAS keywords in Excel but none were applied to the .bib.
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

BIB_PATH = Path("src/metacoupling/data/telecoupling_literature.bib")
XLSX_PATH = Path("D:/Check/journal_research_keywords_extracted.xlsx")
CSV_PATH = Path("tmp/keyword_extraction.csv")
OUT_PATH = Path("tmp/audit_unreplaced.txt")

DELETE_KEYS = {
    # MeSH-junk
    "booth_investigating_2021", "chen_physical_2023", "dasilva_socioeconomic_2021",
    "fastner_telecoupled_2023", "green_linking_2019", "gurney_redefining_2017",
    "sun_importing_2018", "xu_impacts_2020",
    # Placeholder-only `telecoupling`
    "barbieri_food_2022", "chai_telecoupled_2024", "henriksson_measuring_2018",
    "jiren_multi_2022", "scott_remittances_2024", "strecker_kenya_2023",
    "zhang_remittance_2024", "zhao_tightening_2019",
}


def parse_bib_entries(text: str) -> list[dict]:
    """Return list of {key, year, authors, title} for every @article entry."""
    starts = [m for m in re.finditer(r"^@\w+\{([^,]+),", text, re.MULTILINE)]
    entries = []
    for i, m in enumerate(starts):
        key = m.group(1).strip()
        end = starts[i + 1].start() if i + 1 < len(starts) else len(text)
        chunk = text[m.start():end]
        year_m = re.search(r"\byear\s*=\s*\{(\d{4})\}", chunk)
        author_m = re.search(r"\bauthor\s*=\s*\{([^}]*)\}", chunk)
        title_m = re.search(r"\btitle\s*=\s*\{([^}]*)\}", chunk)
        kw_m = re.search(r"\bkeywords\s*=\s*\{([^}]*)\}", chunk)
        entries.append({
            "key": key,
            "year": year_m.group(1) if year_m else "",
            "authors": author_m.group(1).strip() if author_m else "",
            "title": title_m.group(1).strip() if title_m else "",
            "has_keywords": bool(kw_m),
            "keywords": kw_m.group(1).strip() if kw_m else "",
        })
    return entries


def load_user_keywords() -> list[dict]:
    """Return list of {filename, keywords, status} from Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    sheet = wb["All Journal Research"]
    header = [c.value for c in next(sheet.iter_rows())]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        rows.append({
            "filename": d.get("filename") or "",
            "keywords": d.get("keywords") or "",
            "status": d.get("keyword_status") or "",
        })
    return rows


def load_csv_mapping() -> dict[str, str]:
    """Return {filename: bib_key} currently used by integrate_keywords."""
    out = {}
    with open(CSV_PATH, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            out[r["filename"]] = r.get("bib_key", "")
    return out


def normalize(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def title_tokens(s: str) -> set[str]:
    stop = {"the", "a", "an", "of", "and", "in", "on", "for", "to", "by",
            "with", "from", "how", "what", "why", "is", "are", "can", "be",
            "as", "at", "or", "its", "it", "this", "that", "case", "study"}
    return {t for t in normalize(s).split() if len(t) > 2 and t not in stop}


def find_excel_candidates(entry: dict, excel_rows: list[dict],
                          csv_map: dict[str, str]) -> list[dict]:
    """Return Excel rows whose filename looks like a match for this entry.

    Scores by: year match + author surname overlap + title token overlap.
    Also tracks whether that filename is already mapped elsewhere in CSV.
    """
    key = entry["key"]
    year = entry["year"]
    # Extract surnames from `Last, First and Last, First and ...`
    author_field = entry["authors"]
    surnames = set()
    for chunk in re.split(r"\s+and\s+", author_field):
        chunk = chunk.strip()
        if "," in chunk:
            surname = chunk.split(",")[0].strip()
        else:
            surname = chunk.split()[-1] if chunk else ""
        surname = re.sub(r"[^A-Za-z]", "", surname).lower()
        if surname and surname not in {"others", "al", "etal"}:
            surnames.add(surname)
    title_set = title_tokens(entry["title"])

    candidates = []
    for row in excel_rows:
        fn = row["filename"]
        if not fn:
            continue
        fn_norm = normalize(fn)
        # Must match year
        if year and year not in fn:
            continue
        # Require at least one surname appear in filename OR strong title overlap
        surname_hit = any(s in fn_norm for s in surnames)
        fn_title = fn_norm
        # Strip author/year prefix approximately: look at tokens after the year
        after_year = fn_norm.split(year, 1)[1] if year and year in fn_norm else fn_norm
        ft_set = title_tokens(after_year)
        if title_set and ft_set:
            overlap = len(title_set & ft_set) / max(1, min(len(title_set), len(ft_set)))
        else:
            overlap = 0.0
        if not surname_hit and overlap < 0.3:
            continue
        # Score
        score = (1 if surname_hit else 0) + overlap
        candidates.append({
            "filename": fn,
            "keywords": row["keywords"],
            "status": row["status"],
            "score": round(score, 2),
            "overlap": round(overlap, 2),
            "surname_hit": surname_hit,
            "mapped_to_in_csv": csv_map.get(fn, ""),
        })
    candidates.sort(key=lambda c: c["score"], reverse=True)
    return candidates


def main():
    bib_text = BIB_PATH.read_text(encoding="utf-8")
    entries = parse_bib_entries(bib_text)
    by_key = {e["key"]: e for e in entries}

    # PRESERVED = entries without keywords in current .bib, minus those in DELETE_KEYS
    #   (DELETE removed the line, so they also have no keywords now)
    # Actually: DELETE_KEYS intentionally lost their keywords. PRESERVED is
    # defined as "everything else we left alone" — those not in DELETE_KEYS
    # and not touched by REPLACE.
    # To identify PRESERVED: keys not in DELETE_KEYS and not in the CSV's
    # bib_key column (because integration only touched keys we mapped).
    csv_map = load_csv_mapping()
    all_keys = {e["key"] for e in entries}
    excel_rows = load_user_keywords()
    # A key was actually REPLACED only if (a) some Excel row HAS keywords AND
    # (b) that same filename is mapped to the key in the CSV. (Matches what
    # integrate_keywords.py's bibkey_to_user_kw actually gets populated with.)
    replaced_keys: set[str] = set()
    for row in excel_rows:
        if row["keywords"]:
            bk = csv_map.get(row["filename"], "")
            if bk:
                replaced_keys.add(bk)
    preserved_keys = all_keys - DELETE_KEYS - replaced_keys
    preserved_keys = sorted(preserved_keys)
    delete_keys_sorted = sorted(DELETE_KEYS)

    lines: list[str] = []

    def pr(s: str = ""):
        lines.append(s)

    def report(title: str, keys: list[str]):
        pr()
        pr("=" * 78)
        pr(f"{title}  ({len(keys)} entries)")
        pr("=" * 78)
        for k in keys:
            e = by_key.get(k)
            if not e:
                pr(f"\n[{k}]  MISSING FROM BIB")
                continue
            authors = e["authors"][:70]
            pr(f"\n[{k}]  ({e['year']})  {authors}")
            pr(f"   title: {e['title'][:120]}")
            pr(f"   current kw: {(e['keywords'] or '(none)')[:120]}")
            # Only consider candidates that both share the year AND either share
            # a surname with high title overlap, OR are the entry's own mapping.
            cands = find_excel_candidates(e, excel_rows, csv_map)
            # Tighten: require overlap>=0.5 when mapped elsewhere, else surname+overlap>=0.3
            def is_strong(c):
                if c["mapped_to_in_csv"] == k:
                    return True
                return c["surname_hit"] and c["overlap"] >= 0.5
            strong = [c for c in cands if is_strong(c)]
            if not strong:
                pr("   Excel candidates: NONE (no plausibly-matching filename)")
                continue
            for c in strong[:3]:
                flag = ""
                if c["mapped_to_in_csv"] == k:
                    if c["keywords"]:
                        flag = "  (applied in integration)"
                    else:
                        flag = "  <-- Excel has no keywords (confirms empty)"
                elif c["keywords"] and not c["mapped_to_in_csv"]:
                    flag = "  <-- RECOVERABLE (kw in Excel, no CSV mapping)"
                elif c["keywords"] and c["mapped_to_in_csv"] != k:
                    flag = f"  <-- COLLISION (Excel kw mapped to {c['mapped_to_in_csv']})"
                pr(f"   xlsx: score={c['score']} surname={c['surname_hit']} "
                   f"overlap={c['overlap']}  mapped->{c['mapped_to_in_csv'] or '(none)'}"
                   f"{flag}")
                pr(f"          filename: {c['filename'][:110]}")
                pr(f"          kw: {(c['keywords'] or '(none)')[:120]}")
                pr(f"          status: {c['status']}")

    report("16 DELETED -- keywords line removed from .bib", delete_keys_sorted)
    report("20 PRESERVED -- .bib entry untouched by integration", preserved_keys)

    OUT_PATH.parent.mkdir(exist_ok=True)
    OUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT_PATH}  ({len(lines)} lines)")
    # Summary count of actionable flags
    full = "\n".join(lines)
    print(f"  RECOVERABLE flags: {full.count('<-- RECOVERABLE')}")
    print(f"  COLLISION flags:   {full.count('<-- COLLISION')}")
    print(f"  'no keywords' confirms: {full.count('confirms empty')}")


if __name__ == "__main__":
    main()
