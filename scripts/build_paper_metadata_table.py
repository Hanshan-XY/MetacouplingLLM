"""Build a single .xlsx table of paper metadata for all 265 papers.

Source of truth (per user decision):
  - Everything except abstract comes from the .bib (already reconciled
    with user-validated keywords and citation counts).
  - Abstract comes from the full-text markdown in
    D:/Check/_markdown/Journal Articles/Research/.

Output: tmp/paper_metadata.xlsx  (single sheet "Papers")
"""
from __future__ import annotations

import csv
import glob
import re
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

BIB_PATH = Path("src/metacoupling/data/telecoupling_literature.bib")
CSV_PATH = Path("tmp/keyword_extraction.csv")
RESEARCH_DIR = Path(r"D:/Check/_markdown/Journal Articles/Research")
OUT_PATH = Path("tmp/paper_metadata.xlsx")

# Filename -> bib_key overrides for cases where Zotero's filename
# doesn't match the .bib key convention (compound surnames, accents,
# first-name filenames, etc.). Copied verbatim from
# scripts/verify_papers_vs_bib.py where it was curated.
MANUAL_OVERRIDES = {
    "Javier Alcantara-Plazola and de la Barrera - 2021 - Quantification of embedded phosphorus in Mexican agriculture.md": "alcntaraplazola_quantification_2021",
    "Bicudo da Silva et al. - 2019 - Eco-certification protocols as mechanisms to foster sustainable environmental practices in telecoupl.md": "dasilva_eco_2019",
    "Yue et al. - 2018 - Spillover effect offsets the conservation effort in the Amazon.md": "dou_spillover_2018",
    "Du et al. - 2022 - How far are we from possible ideal virtual water transfer Evidence from assessing vulnerability of.md": "du_how_2022_a",
    "ChuangLin and Yufei - 2017 - Analysis of emergy-based metabolic efficiency and environmental pressure on the local coupling and t.md": "fang_analysis_2017",
    "Friis and Nielsen - 2017 - On the System. Boundary Choices, Implications, and Solutions in Telecoupling Land Use Change Researc.md": "friis_system_2017",
    "Lopez-Hoffman et al. - 2017 - Operationalizing the telecoupling framework for migratory species using the spatial subsidies approa.md": "lpezhoffman_operationalizing_2017",
    "de Lucio et al. - 2021 - Resilience as a Moving Target An Evaluation of Last Century Management Strategies in a Dry-Edge Mar.md": "morenofernndez_resilience_2021",
    "Rey and Huettmann - 2020 - Telecoupling analysis of the Patagonian Shelf A new approach to study global seabird-fisheries inte.md": "rayarey_telecoupling_2020",
    "Sondergaard et al. - 2024 - Fragmented sustainability governance of telecoupled flows Brazilian beef exports to China.md": "sndergaard_fragmented_2024",
    "Vila and Arzamendia - 2022 - South American Camelids their values and contributions to people.md": "vil_south_2022",
    "Drakou et al. - 2017 - Drought impacts to water footprints and virtual water transfers of the Central Valley of California.md": "marston_drought_2017",
    "Liu and Pan - 2025 - Unraveling the quantity and sustainability of cross-scale ecosystem service flows A meta-coupling f.md": "liu_unraveling_2025_a",
    "Liu et al. - 2025 - From Plate to Plow How Dietary Shifts Drive Telecoupled Cropland Erosion in China.md": "liu_plate_2025",
    "Li et al. - 2019 - Tightening ecological management facilitates green development in the Qilian Mountains.md": "zhao_tightening_2019",
    "Reis et al. - 2020 - Understanding the Stickiness of Commodity Supply Chains Is Key to Improving Their Sustainability.md": "dosreis_understanding_2020",
    "Zheng and Sun_The short-range and remote analysis of virtual water trade in China.md": "zheng_short_2023",
}


# --- .bib parsing (balanced-brace handles LaTeX like $\mathrm{CO}_{2}$) ---

def extract_field_balanced(chunk: str, field: str) -> str | None:
    m = re.search(rf"\b{field}\s*=\s*\{{", chunk, re.IGNORECASE)
    if not m:
        return None
    i = m.end()
    depth = 1
    while i < len(chunk) and depth > 0:
        c = chunk[i]
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return chunk[m.end():i]
        i += 1
    return None


def parse_bib(text: str) -> dict[str, dict]:
    """bib_key -> dict of all fields we care about."""
    starts = list(re.finditer(r"^@\w+\{([^,]+),", text, re.MULTILINE))
    out = {}
    for i, m in enumerate(starts):
        end = starts[i + 1].start() if i + 1 < len(starts) else len(text)
        chunk = text[m.start():end]
        key = m.group(1).strip()

        title = (extract_field_balanced(chunk, "title") or "").strip()
        author = (extract_field_balanced(chunk, "author") or "").strip()
        year_s = (extract_field_balanced(chunk, "year") or "").strip()
        try:
            year = int(year_s)
        except ValueError:
            year = None
        journal = (extract_field_balanced(chunk, "journal") or "").strip()
        doi = (extract_field_balanced(chunk, "doi") or "").strip()
        keywords = (extract_field_balanced(chunk, "keywords") or "").strip()
        annote = extract_field_balanced(chunk, "annote") or ""
        cite_m = re.search(r"Cited\s*by:\s*(\d+)", annote, re.IGNORECASE)
        cited_by = int(cite_m.group(1)) if cite_m else None

        out[key] = {
            "title": title,
            "authors": author,
            "year": year,
            "journal": journal,
            "doi": doi,
            "keywords": keywords,
            "cited_by": cited_by,
        }
    return out


def load_csv_map() -> dict[str, str]:
    """filename -> bib_key from tmp/keyword_extraction.csv."""
    out = {}
    with open(CSV_PATH, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r.get("bib_key"):
                out[r["filename"]] = r["bib_key"]
    return out


# --- Abstract extraction from markdown ---

def _xml_safe(s):
    """Strip control characters openpyxl/lxml reject for XLSX output."""
    if not isinstance(s, str):
        return s
    # Use openpyxl's own definition plus surrogates.
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    # Also strip unpaired surrogates and noncharacter codepoints.
    s = re.sub(r"[\ud800-\udfff\ufffe\uffff]", "", s)
    return s


def _clean_abstract_block(block: str) -> str:
    """Remove image links, footnote markers; collapse excess blank lines."""
    # strip ![alt](url)
    block = re.sub(r"!\[[^\]]*\]\([^)]*\)", "", block)
    # strip lone footnote markers like [^1]
    block = re.sub(r"\[\^\d+\]", "", block)
    # collapse 3+ consecutive newlines -> 2
    block = re.sub(r"\n{3,}", "\n\n", block)
    return _xml_safe(block.strip())


def _body_until_next_heading(lines: list[str], start_idx: int) -> str:
    """Return text from start_idx until the next markdown heading (#... ).
    Also stops at 4+ consecutive blank lines."""
    out_lines: list[str] = []
    blank_run = 0
    for ln in lines[start_idx:]:
        if re.match(r"^#{1,6}\s", ln):
            break
        if ln.strip() == "":
            blank_run += 1
            if blank_run >= 4:
                break
        else:
            blank_run = 0
        out_lines.append(ln)
    return "\n".join(out_lines).strip()


def _heading_norm(ln: str) -> str | None:
    """For a markdown heading line, return its content lowercased with
    all internal whitespace removed. Returns None if not a heading."""
    m = re.match(r"^(#{1,6})\s+(.*?)\s*$", ln)
    if not m:
        return None
    return re.sub(r"\s+", "", m.group(2)).lower()


# Headings that contain "abstract" but are NOT the actual abstract
# (graphical / visual / video abstracts).
_FAKE_ABSTRACT_HEADINGS = {
    "graphicalabstract", "visualabstract", "videoabstract",
    "graphicalabstract.", "plainlanguagesummary",
}


def extract_abstract(md_text: str) -> tuple[str, str]:
    """Return (abstract, status). Status in {'extracted', 'not_found'}."""
    lines = md_text.splitlines()
    heading = [_heading_norm(ln) for ln in lines]

    # Pattern 1: heading whose normalized content is exactly "abstract"
    # Matches "# Abstract", "## ABSTRACT", "# A B S T R A C T", etc.
    for i, h in enumerate(heading):
        if h == "abstract":
            body = _body_until_next_heading(lines, i + 1)
            cleaned = _clean_abstract_block(body)
            if cleaned:
                return cleaned, "extracted"

    # Pattern 2: inline "Abstract" as a paragraph prefix.
    # Accept separators: ':' / '.' / tab / 2+ spaces / single space
    # (single-space requires content >=80 chars to avoid matching stray
    # sentences like "Abstract reasoning is ...").
    for i, ln in enumerate(lines):
        if ln.lstrip().startswith("#"):
            continue
        m = re.match(
            r"^\s*Abstract(?:\s*[:\.]\s+|\s{2,}|\t+|\s+)(.+)$",
            ln, re.IGNORECASE,
        )
        if not m:
            continue
        first = m.group(1).strip()
        # Require the start-of-abstract line to be substantive.
        if len(first) < 80:
            continue
        rest = _body_until_next_heading(lines, i + 1)
        body = (first + "\n" + rest).strip() if first else rest
        cleaned = _clean_abstract_block(body)
        if cleaned:
            return cleaned, "extracted"

    # Pattern 3: heading whose normalized content contains "abstract" but
    # isn't a graphical/visual abstract placeholder.
    for i, h in enumerate(heading):
        if h and "abstract" in h and h not in _FAKE_ABSTRACT_HEADINGS:
            if h.startswith("graphicalabstract") or h.startswith("visualabstract"):
                continue
            body = _body_until_next_heading(lines, i + 1)
            cleaned = _clean_abstract_block(body)
            if cleaned:
                return cleaned, "extracted"

    # Pattern 4: heading "# Summary" (used by J. Industrial Ecology etc.
    # in lieu of "Abstract"). Only fires when no abstract heading was
    # found — we're already past Patterns 1-3.
    for i, h in enumerate(heading):
        if h == "summary":
            body = _body_until_next_heading(lines, i + 1)
            cleaned = _clean_abstract_block(body)
            if cleaned:
                return cleaned, "extracted"

    # Pattern 5: inline "Non-technical summary. ..." / "Technical summary. ..."
    # / "Summary. ..." as paragraph prefix (Cambridge/Global Sustainability).
    summary_pieces: list[str] = []
    for i, ln in enumerate(lines):
        if ln.lstrip().startswith("#"):
            continue
        m = re.match(
            r"^\s*(?:Non-technical|Technical)?\s*Summary\s*[:\.]\s+(.{80,})$",
            ln, re.IGNORECASE,
        )
        if m:
            summary_pieces.append(m.group(1).strip())
    if summary_pieces:
        joined = "\n\n".join(summary_pieces)
        cleaned = _clean_abstract_block(joined)
        if cleaned:
            return cleaned, "extracted"

    # Pattern 6: Chinese-language abstract marker "摘要 ..." inline.
    for i, ln in enumerate(lines):
        m = re.match(r"^\s*摘\s*要\s+(.{40,})$", ln)
        if m:
            body = m.group(1).strip()
            cleaned = _clean_abstract_block(body)
            if cleaned:
                return cleaned, "extracted"

    return "", "not_found"


# --- XLSX writer ---

COLUMNS = [
    ("bib_key", 30),
    ("title", 60),
    ("authors", 40),
    ("year", 8),
    ("journal", 30),
    ("doi", 32),
    ("keywords", 50),
    ("cited_by", 10),
    ("abstract", 120),
    ("abstract_status", 16),
    ("markdown_filename", 60),
]


def write_xlsx(rows: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers"

    # Header
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (name, _) in enumerate(COLUMNS, start=1):
            value = row.get(name)
            if isinstance(value, str):
                value = _xml_safe(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    bib = parse_bib(BIB_PATH.read_text(encoding="utf-8"))
    print(f".bib entries:        {len(bib)}")

    fname_to_key = load_csv_map()
    fname_to_key.update(MANUAL_OVERRIDES)

    # Inverse: bib_key -> markdown file path (if any)
    key_to_md: dict[str, Path] = {}
    md_files = sorted(glob.glob(str(RESEARCH_DIR / "**/*.md"), recursive=True))
    print(f"Markdown files found: {len(md_files)}")
    for md in md_files:
        name = Path(md).name
        key = fname_to_key.get(name)
        if key and key not in key_to_md:
            key_to_md[key] = Path(md)

    stats = {"extracted": 0, "not_found": 0, "no_markdown": 0}
    rows: list[dict] = []
    for key in sorted(bib):
        meta = bib[key]
        md_path = key_to_md.get(key)
        abstract = ""
        abs_status = "no_markdown"
        if md_path:
            md_text = md_path.read_text(encoding="utf-8", errors="replace")
            abstract, abs_status = extract_abstract(md_text)
        stats[abs_status] += 1
        rows.append({
            "bib_key": key,
            "title": meta["title"],
            "authors": meta["authors"],
            "year": meta["year"],
            "journal": meta["journal"],
            "doi": meta["doi"],
            "keywords": meta["keywords"],
            "cited_by": meta["cited_by"],
            "abstract": abstract,
            "abstract_status": abs_status,
            "markdown_filename": md_path.name if md_path else "",
        })

    write_xlsx(rows, OUT_PATH)
    print(f"Markdown matched:    {sum(1 for r in rows if r['markdown_filename'])}")
    print(f"Abstracts extracted: {stats['extracted']}")
    print(f"Abstracts not found: {stats['not_found']}")
    print(f"No markdown:         {stats['no_markdown']}")
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
