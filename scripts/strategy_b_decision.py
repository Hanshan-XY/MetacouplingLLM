"""Apply Strategy B (ship CC-BY + SA + NC + NC-SA; skip ND/NC-ND/closed/bronze
/green-preprint) and write tmp/strategy_b_decision.xlsx with two sheets:
  - 'Journal Articles' (420 papers)
  - 'Book Chapters'    (25 papers)

Each row: filename, title, year, doi, oa_status, license, ship, reason.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

IN_PATH = Path("tmp/oa_status_report.xlsx")
OUT_PATH = Path("tmp/strategy_b_decision.xlsx")

# Strategy B: permissive (CC-BY family without ND) + NC variants without ND
SHIP_LICENSES = {"cc-by", "cc-by-sa", "cc-by-nc", "cc-by-nc-sa",
                 "cc0", "pd", "public-domain"}


def xml_safe(s):
    if not isinstance(s, str):
        return s
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return re.sub(r"[\ud800-\udfff￾￿]", "", s)


def decide(license_: str, oa_status: str, doi: str, version: str) -> tuple[str, str]:
    license_ = (license_ or "").lower().strip()
    oa_status = (oa_status or "").lower().strip()
    version = (version or "").lower().strip()

    if not doi:
        return "No", "No DOI"
    if license_ in SHIP_LICENSES:
        # Ensure status is OA (gold/hybrid/green publishedVersion)
        if oa_status in ("gold", "hybrid"):
            return "Yes", f"{oa_status} OA, {license_}"
        if oa_status == "green":
            if version == "publishedversion":
                return "Yes", f"green OA published version, {license_}"
            return "No", f"green OA but only {version or 'unknown'} version available"
        return "No", f"OA status {oa_status or 'unknown'} unclear despite license {license_}"
    if license_.endswith("-nd"):
        return "No", f"ND license ({license_}) prohibits derivative chunks/embeddings"
    if oa_status == "bronze":
        return "No", "bronze OA, free to read but no license to redistribute"
    if oa_status == "closed":
        return "No", "closed access (subscription only)"
    if oa_status == "green":
        return "No", f"green OA without permissive license ({license_ or 'none'})"
    if license_:
        return "No", f"non-permissive license ({license_})"
    return "No", f"no license info, oa_status={oa_status or 'unknown'}"


COLUMNS = [
    ("ship", 8),
    ("reason", 50),
    ("filename", 60),
    ("title", 60),
    ("year", 8),
    ("doi", 32),
    ("oa_status", 10),
    ("license", 14),
    ("category", 30),
]


def write_sheet(ws, rows: list[dict]) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    yes_fill = PatternFill("solid", fgColor="E2EFDA")  # light green
    no_fill = PatternFill("solid", fgColor="FCE4D6")   # light orange
    header_font = Font(bold=True)
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = width

    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(rows, start=2):
        ship = row.get("ship", "")
        for c_idx, (name, _) in enumerate(COLUMNS, start=1):
            v = row.get(name)
            if isinstance(v, str):
                v = xml_safe(v)
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = wrap
            if name == "ship":
                cell.fill = yes_fill if ship == "Yes" else no_fill
                cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def main() -> None:
    wb_in = openpyxl.load_workbook(IN_PATH, read_only=True, data_only=True)
    sh_in = wb_in["OA Status"]
    rows_in = list(sh_in.iter_rows(values_only=True))
    idx = {n: i for i, n in enumerate(rows_in[0])}

    book = []
    journal = []
    counts = {"book_yes": 0, "book_no": 0, "journal_yes": 0, "journal_no": 0}
    for r in rows_in[1:]:
        category = r[idx["category"]] or ""
        ship, reason = decide(
            r[idx["license"]] or "",
            r[idx["oa_status"]] or "",
            r[idx["doi"]] or "",
            r[idx["version"]] or "",
        )
        out = {
            "ship": ship,
            "reason": reason,
            "filename": r[idx["filename"]] or "",
            "title": r[idx["title"]] or "",
            "year": r[idx["year"]],
            "doi": r[idx["doi"]] or "",
            "oa_status": r[idx["oa_status"]] or "",
            "license": r[idx["license"]] or "",
            "category": category,
        }
        if category == "Book Chapters":
            book.append(out)
            counts["book_yes" if ship == "Yes" else "book_no"] += 1
        elif category.startswith("Journal Articles"):
            journal.append(out)
            counts["journal_yes" if ship == "Yes" else "journal_no"] += 1

    # Sort: Yes first then alphabetical filename
    sort_key = lambda r: (0 if r["ship"] == "Yes" else 1, r["filename"].lower())
    book.sort(key=sort_key)
    journal.sort(key=sort_key)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    ws1 = wb_out.create_sheet("Journal Articles")
    write_sheet(ws1, journal)
    ws2 = wb_out.create_sheet("Book Chapters")
    write_sheet(ws2, book)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(OUT_PATH)

    print(f"Strategy B decisions written to {OUT_PATH}")
    print()
    print(f"Journal Articles: {len(journal)} total")
    print(f"  Yes (ship): {counts['journal_yes']}")
    print(f"  No  (skip): {counts['journal_no']}")
    print()
    print(f"Book Chapters: {len(book)} total")
    print(f"  Yes (ship): {counts['book_yes']}")
    print(f"  No  (skip): {counts['book_no']}")
    print()
    print(f"Total ship: {counts['journal_yes'] + counts['book_yes']} of "
          f"{len(journal) + len(book)}")


if __name__ == "__main__":
    main()
