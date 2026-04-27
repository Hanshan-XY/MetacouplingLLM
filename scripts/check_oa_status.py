"""Check OA status of every PDF in D:/Allpdf/Journal Articles/ +
D:/Allpdf/Book Chapters/ via the Unpaywall API.

Output: tmp/oa_status_report.xlsx (single sheet 'OA Status', 445 rows)

Caches:
  - tmp/doi_inventory.json     (PDF -> DOI; speeds up re-runs)
  - tmp/unpaywall_cache.json   (DOI -> Unpaywall response)

Both are safe to delete at any time; the script will rebuild them.
"""
from __future__ import annotations

import csv
import json
import re
import time
from pathlib import Path
from typing import Optional

import fitz  # PyMuPDF
import openpyxl
import requests
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PDF_ROOT = Path(r"D:\Allpdf")
SCAN_FOLDERS = ["Journal Articles", "Book Chapters"]
BIB_PATH = Path("src/metacoupling/data/telecoupling_literature.bib")
EXISTING_CSV = Path("paper_citation_counts.csv")
DOI_CACHE = Path("tmp/doi_inventory.json")
UNPAYWALL_CACHE = Path("tmp/unpaywall_cache.json")
OUT_PATH = Path("tmp/oa_status_report.xlsx")

# DOI regex from build_rag.py:46
DOI_RE = re.compile(r"\b(10\.\d{4,}/[^\s\]>\"',;]+)")

UNPAYWALL_URL = "https://api.unpaywall.org/v2/{doi}?email=anonymous@example.org"
UNPAYWALL_TIMEOUT = 20

# Filename -> bib_key overrides (copied from verify_papers_vs_bib.py)
MANUAL_OVERRIDES = {
    "Javier Alcantara-Plazola and de la Barrera - 2021 - Quantification of embedded phosphorus in Mexican agriculture.md": "alcntaraplazola_quantification_2021",
    "Bicudo da Silva et al. - 2019 - Eco-certification protocols as mechanisms to foster sustainable environmental practices in telecoupl.md": "dasilva_eco_2019",
    "Yue et al. - 2018 - Spillover effect offsets the conservation effort in the Amazon.md": "dou_spillover_2018",
    "Du et al. - 2022 - How far are we from possible ideal virtual water transfer Evidence from assessing vulnerability of.md": "du_how_2022_a",
    "ChuangLin and Yufei - 2017 - Analysis of emergy-based metabolic efficiency and environmental pressure on the local coupling and t.md": "fang_analysis_2017",
    "Friis and Nielsen - 2017 - On the System. Boundary Choices, Implications, and Solutions in Telecoupling Land Use Change Researc.md": "friis_system_2017",
    "Lopez-Hoffman et al. - 2017 - Operationalizing the telecoupling framework for migratory species using the spatial subsidies approa.md": "lpezhoffman_operationalizing_2017",
    "de Lucio et al. - 2021 - Resilience as a Moving Target An Evaluation of Last Century Management Strategies in a Dry-Edge Mar.md": "morenofernndez_resilience_2021",
    "Rey and Huettmann - 2020 - Telecoupling analysis of the Patagonian Shelf A new approach to study global seabird-fisheries inte.md": "rayarey_telecoupling_2020",
    "Sondergaard et al. - 2024 - Fragmented sustainability governance of telecoupled flows Brazilian beef exports to China.md": "sndergaard_fragmented_2024",
    "Vila and Arzamendia - 2022 - South American Camelids their values and contributions to people.md": "vil_south_2022",
    "Drakou et al. - 2017 - Drought impacts to water footprints and virtual water transfers of the Central Valley of California.md": "marston_drought_2017",
    "Liu and Pan - 2025 - Unraveling the quantity and sustainability of cross-scale ecosystem service flows A meta-coupling f.md": "liu_unraveling_2025_a",
    "Liu et al. - 2025 - From Plate to Plow How Dietary Shifts Drive Telecoupled Cropland Erosion in China.md": "liu_plate_2025",
    "Li et al. - 2019 - Tightening ecological management facilitates green development in the Qilian Mountains.md": "zhao_tightening_2019",
    "Reis et al. - 2020 - Understanding the Stickiness of Commodity Supply Chains Is Key to Improving Their Sustainability.md": "dosreis_understanding_2020",
    "Zheng and Sun_The short-range and remote analysis of virtual water trade in China.md": "zheng_short_2023",
}
# Strip the .md extension for matching against PDF stems (which have no ext)
MANUAL_OVERRIDES_STEM = {Path(k).stem: v for k, v in MANUAL_OVERRIDES.items()}


# ---------------------------------------------------------------------------
# Helpers reused from sister scripts
# ---------------------------------------------------------------------------

def extract_field_balanced(chunk: str, field: str) -> Optional[str]:
    """From verify_papers_vs_bib.py — handles LaTeX braces."""
    m = re.search(rf"\b{field}\s*=\s*\{{", chunk, re.IGNORECASE)
    if not m:
        return None
    i = m.end()
    depth = 1
    while i < len(chunk) and depth > 0:
        c = chunk[i]
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return chunk[m.end():i]
        i += 1
    return None


def parse_bib(text: str) -> dict[str, dict]:
    """bib_key -> {title, year, doi}."""
    starts = list(re.finditer(r"^@\w+\{([^,]+),", text, re.MULTILINE))
    out = {}
    for i, m in enumerate(starts):
        end = starts[i + 1].start() if i + 1 < len(starts) else len(text)
        chunk = text[m.start():end]
        key = m.group(1).strip()
        out[key] = {
            "title": (extract_field_balanced(chunk, "title") or "").strip(),
            "year": (extract_field_balanced(chunk, "year") or "").strip(),
            "doi": (extract_field_balanced(chunk, "doi") or "").strip(),
        }
    return out


def extract_doi_from_pdf(pdf_path: Path) -> Optional[str]:
    """From build_rag.py:83-99 — scan first 3 pages for DOI."""
    try:
        doc = fitz.open(str(pdf_path))
        text = ""
        for i in range(min(3, doc.page_count)):
            text += doc[i].get_text()
        doc.close()
        m = DOI_RE.search(text)
        if m:
            return m.group(1).rstrip(".,;:)")
    except Exception:
        pass
    return None


def xml_safe(s):
    if not isinstance(s, str):
        return s
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return re.sub(r"[\ud800-\udfff￾￿]", "", s)


# ---------------------------------------------------------------------------
# Phase 1: DOI inventory
# ---------------------------------------------------------------------------

def build_doi_inventory(
    pdfs: list[Path],
    bib: dict[str, dict],
    pdf_to_bibkey: dict[str, str],
    cache: dict[str, dict],
) -> dict[str, dict]:
    """For each PDF, find its DOI. Returns {pdf_path_str: {doi, doi_source, bib_key}}."""
    inventory = {}
    new_count = 0
    for pdf in pdfs:
        key = str(pdf)
        if key in cache:
            inventory[key] = cache[key]
            continue

        bib_key = pdf_to_bibkey.get(pdf.stem) or pdf_to_bibkey.get(pdf.name)
        # Try .bib first
        if bib_key and bib.get(bib_key, {}).get("doi"):
            entry = {
                "doi": bib[bib_key]["doi"],
                "doi_source": "bib",
                "bib_key": bib_key,
            }
        else:
            # PDF scan
            doi = extract_doi_from_pdf(pdf)
            entry = {
                "doi": doi or "",
                "doi_source": "pdf_scan" if doi else "not_found",
                "bib_key": bib_key or "",
            }
        inventory[key] = entry
        new_count += 1
        if new_count % 25 == 0:
            print(f"  scanned {new_count} PDFs...")

    return inventory


# ---------------------------------------------------------------------------
# Phase 2: Unpaywall lookups
# ---------------------------------------------------------------------------

def unpaywall_lookup(doi: str, cache: dict, throttle_s: float = 0.2) -> dict:
    if doi in cache:
        return cache[doi]
    url = UNPAYWALL_URL.format(doi=doi)
    try:
        r = requests.get(url, timeout=UNPAYWALL_TIMEOUT,
                         headers={"User-Agent": "metacoupling-rag/1.0"})
        if r.status_code == 200:
            data = r.json()
            result = {
                "ok": True,
                "is_oa": bool(data.get("is_oa")),
                "oa_status": data.get("oa_status") or "",
                "best_loc": data.get("best_oa_location") or {},
                "raw_status_code": r.status_code,
            }
        elif r.status_code == 404:
            result = {"ok": False, "reason": "404_not_in_unpaywall",
                      "raw_status_code": 404}
        else:
            result = {"ok": False, "reason": f"http_{r.status_code}",
                      "raw_status_code": r.status_code}
    except Exception as e:
        result = {"ok": False, "reason": f"exception:{type(e).__name__}",
                  "raw_status_code": None}
    cache[doi] = result
    time.sleep(throttle_s)
    return result


# ---------------------------------------------------------------------------
# Phase 3: Redistribution-safety classification
# ---------------------------------------------------------------------------

_PERMISSIVE = {"cc-by", "cc-by-sa", "cc0", "pd", "public-domain"}
_NC_OR_ND = ("cc-by-nc", "cc-by-nd", "cc-by-nc-sa", "cc-by-nc-nd")


def classify_redistribution(unpaywall_entry: dict) -> tuple[str, str]:
    """Returns (yes/maybe/no, notes).

    For a published Python package that ships full text inside Papers.zip:
      - Yes:   permissive (CC-BY, CC-BY-SA, CC0, public domain)
      - Maybe: NC or ND restrictions — depends on package license
               (NC ok if package non-commercial; ND blocks chunking/embeddings)
      - No:    bronze (no license), closed, or unrecognized license
    """
    if not unpaywall_entry.get("ok"):
        return "No", f"Unpaywall lookup failed: {unpaywall_entry.get('reason', 'unknown')}"
    status = (unpaywall_entry.get("oa_status") or "").lower()
    loc = unpaywall_entry.get("best_loc") or {}
    license_ = (loc.get("license") or "").lower()
    version = (loc.get("version") or "").lower()

    # Gold / hybrid: license-driven decision
    if status in ("gold", "hybrid"):
        if license_ in _PERMISSIVE:
            return "Yes", f"{status} OA, permissive license ({license_})"
        if license_.startswith(_NC_OR_ND):
            return "Maybe", (f"{status} OA but {license_} — NC/ND restricts "
                              "commercial reuse and/or derivatives")
        if license_:
            return "Maybe", f"{status} OA, non-standard license ({license_})"
        return "No", f"{status} OA but no license info"

    # Green: published version with permissive CC is okay; otherwise risky
    if status == "green":
        if version == "publishedversion" and license_ in _PERMISSIVE:
            return "Yes", f"Green OA publishedVersion with {license_}"
        if license_ in _PERMISSIVE:
            return "Maybe", f"Green OA accepted/preprint version with {license_}"
        if license_.startswith(_NC_OR_ND):
            return "Maybe", f"Green OA with {license_} — NC/ND restrictions"
        return "No", f"Green OA but version={version or 'unknown'}, license={license_ or 'none'}"

    # Bronze / closed / etc.
    if status == "bronze":
        return "No", "Bronze OA: free to read but no redistribution license"
    if status == "closed":
        return "No", "Closed access (subscription only)"
    return "No", f"Unknown OA status: {status or 'empty'}"


# ---------------------------------------------------------------------------
# Phase 4: XLSX writer
# ---------------------------------------------------------------------------

COLUMNS = [
    ("category", 30),
    ("filename", 60),
    ("bib_key", 30),
    ("title", 60),
    ("year", 8),
    ("doi", 32),
    ("doi_source", 12),
    ("is_oa", 8),
    ("oa_status", 10),
    ("license", 14),
    ("version", 18),
    ("oa_url", 45),
    ("can_redistribute_full_text", 10),
    ("notes", 50),
    ("existing_oa_label", 14),
    ("label_changed", 14),
]


def write_xlsx(rows: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OA Status"

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = width

    wrap = Alignment(wrap_text=True, vertical="top")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (name, _) in enumerate(COLUMNS, start=1):
            v = row.get(name)
            if isinstance(v, str):
                v = xml_safe(v)
            ws.cell(row=r_idx, column=c_idx, value=v).alignment = wrap

    ws.freeze_panes = "A2"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def load_existing_csv() -> dict[str, str]:
    """{filename: existing_oa_label}. Filenames in CSV use .pdf extension."""
    if not EXISTING_CSV.exists():
        return {}
    out = {}
    with open(EXISTING_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            fn = r.get("filename", "")
            if fn:
                out[fn] = (r.get("oa_status") or "").strip()
    return out


def main() -> None:
    # Collect PDFs
    pdfs: list[Path] = []
    for sub in SCAN_FOLDERS:
        d = PDF_ROOT / sub
        if d.exists():
            pdfs.extend(sorted(d.rglob("*.pdf")))
    print(f"PDFs to scan: {len(pdfs)}")

    # Load .bib
    bib = parse_bib(BIB_PATH.read_text(encoding="utf-8"))
    pdf_to_bibkey: dict[str, str] = {}
    # Default mapping: pdf stem == bib key's title-derived filename. We only
    # have the manual overrides for filename↔key mismatches; for the rest the
    # pdf's stem won't directly match the bib key, so we look up by both
    # parts when possible. Easier: build {md_stem: bib_key} from the markdown
    # tree (which we already mapped 1:1) — but for OA we only need the DOI,
    # so the simpler strategy is: try MANUAL_OVERRIDES (handles 17 quirks),
    # then fall back to PDF scan.
    pdf_to_bibkey.update(MANUAL_OVERRIDES_STEM)
    # Plus any direct bib-key matches (pdf stem == bib key) — rare but possible
    for k in bib:
        pdf_to_bibkey.setdefault(k, k)

    # Load caches
    DOI_CACHE.parent.mkdir(parents=True, exist_ok=True)
    doi_cache = json.loads(DOI_CACHE.read_text(encoding="utf-8")) if DOI_CACHE.exists() else {}
    unpaywall_cache = json.loads(UNPAYWALL_CACHE.read_text(encoding="utf-8")) if UNPAYWALL_CACHE.exists() else {}

    # Phase 1: DOI inventory
    print("\nPhase 1: DOI inventory")
    inventory = build_doi_inventory(pdfs, bib, pdf_to_bibkey, doi_cache)
    DOI_CACHE.write_text(json.dumps(inventory, indent=2, ensure_ascii=False),
                         encoding="utf-8")
    sources = {}
    for v in inventory.values():
        sources[v["doi_source"]] = sources.get(v["doi_source"], 0) + 1
    print(f"  DOI sources: {sources}")

    # Phase 2: Unpaywall
    print("\nPhase 2: Unpaywall lookup")
    unique_dois = sorted({v["doi"] for v in inventory.values() if v["doi"]})
    print(f"  Unique DOIs: {len(unique_dois)}")
    new_calls = 0
    for i, doi in enumerate(unique_dois, 1):
        if doi in unpaywall_cache:
            continue
        unpaywall_lookup(doi, unpaywall_cache, throttle_s=0.2)
        new_calls += 1
        if new_calls % 20 == 0:
            print(f"  ...{new_calls} new lookups")
            UNPAYWALL_CACHE.write_text(
                json.dumps(unpaywall_cache, indent=2, ensure_ascii=False),
                encoding="utf-8")
    UNPAYWALL_CACHE.write_text(
        json.dumps(unpaywall_cache, indent=2, ensure_ascii=False),
        encoding="utf-8")
    print(f"  cached: {len(unpaywall_cache)}  new: {new_calls}")

    # Phase 3 + 4: build rows + xlsx
    print("\nPhase 3+4: classify and write xlsx")
    existing_csv = load_existing_csv()

    rows = []
    counts_status: dict[str, int] = {}
    counts_redist = {"Yes": 0, "Maybe": 0, "No": 0}
    not_found = 0
    label_changed = 0

    for pdf in pdfs:
        inv = inventory[str(pdf)]
        doi = inv["doi"]
        bib_key = inv["bib_key"]
        bib_entry = bib.get(bib_key, {}) if bib_key else {}
        rel = pdf.relative_to(PDF_ROOT)
        category = str(rel.parent)

        upw = unpaywall_cache.get(doi, {}) if doi else {}
        loc = upw.get("best_loc") or {}
        license_ = loc.get("license") or ""
        version = loc.get("version") or ""
        oa_url = loc.get("url_for_pdf") or loc.get("url") or ""
        oa_status_str = (upw.get("oa_status") or "").lower() if upw else ""
        if not doi:
            oa_status_str = ""
        counts_status[oa_status_str or "no_doi"] = counts_status.get(
            oa_status_str or "no_doi", 0) + 1
        if inv["doi_source"] == "not_found":
            not_found += 1

        if doi and upw:
            redist, notes = classify_redistribution(upw)
        elif not doi:
            redist, notes = "No", "No DOI found in PDF or .bib"
        else:
            redist, notes = "No", "Unpaywall lookup did not return data"
        counts_redist[redist] += 1

        existing_label = existing_csv.get(pdf.name, "")
        # Map our new label to the binary scheme for diffing
        if redist == "Yes":
            new_label_simple = "OA"
        else:
            new_label_simple = "non-OA"
        changed = ""
        if existing_label and existing_label != new_label_simple:
            changed = f"{existing_label} -> {new_label_simple}"
            label_changed += 1

        rows.append({
            "category": category,
            "filename": pdf.name,
            "bib_key": bib_key,
            "title": bib_entry.get("title", ""),
            "year": int(bib_entry["year"]) if bib_entry.get("year", "").isdigit() else None,
            "doi": doi,
            "doi_source": inv["doi_source"],
            "is_oa": bool(upw.get("is_oa")) if upw.get("ok") else "",
            "oa_status": oa_status_str,
            "license": license_,
            "version": version,
            "oa_url": oa_url,
            "can_redistribute_full_text": redist,
            "notes": notes,
            "existing_oa_label": existing_label,
            "label_changed": changed,
        })

    rows.sort(key=lambda r: (r["category"], r["filename"]))
    write_xlsx(rows, OUT_PATH)

    print()
    print(f"PDFs scanned:        {len(pdfs)}")
    print(f"DOIs found:          {len(pdfs) - not_found}  (not_found: {not_found})")
    print(f"Unpaywall responses: {sum(1 for r in rows if r['oa_status'])}")
    print(f"  by oa_status:")
    for k in sorted(counts_status):
        print(f"    {k:12s}: {counts_status[k]}")
    print(f"can_redistribute_full_text:")
    for k, n in counts_redist.items():
        print(f"  {k:6s}: {n}")
    print(f"label_changed (vs paper_citation_counts.csv): {label_changed}")
    print(f"\nWrote {OUT_PATH}")


if __name__ == "__main__":
    main()
