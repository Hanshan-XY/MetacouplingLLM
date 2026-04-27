"""Conservative copyright audit of the Strategy B 'Yes' list.

Performs five checks:
  A. DOI duplicates — multiple PDFs claiming the same DOI is a red flag.
  B. License consistency — re-verify each Yes paper against
     unpaywall_cache.json (in case the xlsx got out of sync).
  C. License-version sanity — flag cases where license=cc-by but
     version=submittedVersion (preprint license may differ from
     publisher's final article).
  D. Strict-safe vs NC-conditional split — separate truly permissive
     (CC-BY, CC-BY-SA, CC0, public-domain) from NC variants whose
     legality depends on the host package being non-commercial.
  E. Per-license inventory — exhaustive list per license.

Outputs:
  tmp/copyright_audit.txt           — human-readable report
  tmp/strategy_b_safe_strict.xlsx    — only the truly permissive papers
"""
from __future__ import annotations

import json
import re
import sys
import io
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DECISION_XLSX = Path("tmp/strategy_b_decision.xlsx")
UNPAYWALL_CACHE = Path("tmp/unpaywall_cache.json")
DOI_INVENTORY = Path("tmp/doi_inventory.json")
OUT_TXT = Path("tmp/copyright_audit.txt")
OUT_STRICT_XLSX = Path("tmp/strategy_b_safe_strict.xlsx")

STRICT_SAFE = {"cc-by", "cc-by-sa", "cc0", "pd", "public-domain"}
NC_VARIANTS = {"cc-by-nc", "cc-by-nc-sa"}


def safe(s):
    return (s or "").encode("ascii", "replace").decode("ascii")


def xml_safe(s):
    if not isinstance(s, str):
        return s
    s = ILLEGAL_CHARACTERS_RE.sub("", s)
    return re.sub(r"[\ud800-\udfff￾￿]", "", s)


def load_decision_yes() -> list[dict]:
    """Read all 'Yes' rows from both sheets of strategy_b_decision.xlsx."""
    wb = openpyxl.load_workbook(DECISION_XLSX, read_only=True, data_only=True)
    out = []
    for sheet_name in ["Journal Articles", "Book Chapters"]:
        sh = wb[sheet_name]
        rows = list(sh.iter_rows(values_only=True))
        idx = {n: i for i, n in enumerate(rows[0])}
        for r in rows[1:]:
            if r[idx["ship"]] == "Yes":
                out.append({
                    "sheet": sheet_name,
                    "filename": r[idx["filename"]] or "",
                    "title": r[idx["title"]] or "",
                    "year": r[idx["year"]],
                    "doi": r[idx["doi"]] or "",
                    "oa_status": r[idx["oa_status"]] or "",
                    "license": (r[idx["license"]] or "").lower(),
                    "category": r[idx["category"]] or "",
                })
    return out


def main() -> None:
    yes_rows = load_decision_yes()
    unpaywall = json.loads(UNPAYWALL_CACHE.read_text(encoding="utf-8"))
    print(f"Yes rows under Strategy B: {len(yes_rows)}")

    lines: list[str] = []
    pr = lines.append

    pr("=" * 78)
    pr("COPYRIGHT AUDIT — Strategy B 'Yes' list")
    pr("=" * 78)
    pr(f"Total papers marked Yes: {len(yes_rows)}")
    pr("")

    # ----- Check A: DOI duplicates -----
    pr("=" * 78)
    pr("A. DOI DUPLICATE CHECK")
    pr("=" * 78)
    by_doi = defaultdict(list)
    for r in yes_rows:
        if r["doi"]:
            by_doi[r["doi"].lower()].append(r)
    dupes = {d: rs for d, rs in by_doi.items() if len(rs) > 1}
    pr(f"Unique DOIs in Yes list: {len(by_doi)}")
    pr(f"DOIs claimed by 2+ PDFs: {len(dupes)}")
    if dupes:
        pr("FLAG: a duplicate usually means the PDF DOI extractor picked up")
        pr("a citation/reference DOI instead of the paper's own DOI.")
        pr("")
        for doi, rs in sorted(dupes.items()):
            pr(f"  DOI: {doi}")
            for r in rs:
                pr(f"    [{r['sheet']}] {safe(r['filename'])[:90]}")
    else:
        pr("  OK — every Yes paper has a unique DOI.")
    pr("")

    # ----- Check B: License consistency -----
    pr("=" * 78)
    pr("B. LICENSE CONSISTENCY CHECK")
    pr("(re-verify each Yes paper's license against unpaywall_cache.json)")
    pr("=" * 78)
    mismatches = []
    no_cache = []
    for r in yes_rows:
        upw = unpaywall.get(r["doi"])
        if not upw:
            no_cache.append(r)
            continue
        if not upw.get("ok"):
            no_cache.append(r)
            continue
        cache_lic = ((upw.get("best_loc") or {}).get("license") or "").lower()
        if cache_lic != r["license"]:
            mismatches.append((r, cache_lic))
    pr(f"Papers with no Unpaywall cache entry: {len(no_cache)}")
    pr(f"License-string mismatches (xlsx vs cache): {len(mismatches)}")
    if mismatches:
        for r, cache_lic in mismatches[:20]:
            pr(f"  xlsx={r['license']!r}  cache={cache_lic!r}  "
               f"{safe(r['filename'])[:80]}")
    else:
        pr("  OK — every Yes paper's license matches its Unpaywall record.")
    pr("")

    # ----- Check C: License-version sanity -----
    pr("=" * 78)
    pr("C. LICENSE × VERSION SANITY CHECK")
    pr("=" * 78)
    suspect = []
    for r in yes_rows:
        upw = unpaywall.get(r["doi"], {})
        loc = upw.get("best_loc") or {}
        version = (loc.get("version") or "").lower()
        host = (loc.get("host_type") or "").lower()
        if version != "publishedversion":
            suspect.append((r, version, host, loc.get("url") or ""))
    pr(f"Yes papers whose Unpaywall best_oa_location is NOT publishedVersion: "
       f"{len(suspect)}")
    pr("(These have a permissive license on the *self-archived/preprint* copy,")
    pr(" but the journal's published version may carry a different license.")
    pr(" Worth a manual check — bundling the publisher PDF could violate the")
    pr(" publisher's actual terms, even if the preprint is CC-BY.)")
    pr("")
    for r, version, host, url in suspect[:30]:
        pr(f"  [{r['license']:14}] [version={version or '-':17}] "
           f"[host={host or '-':10}] {safe(r['filename'])[:70]}")
        if url:
            pr(f"     OA url: {url[:120]}")
    if len(suspect) > 30:
        pr(f"  ... and {len(suspect) - 30} more")
    pr("")

    # ----- Check D: Strict-safe vs NC-conditional -----
    pr("=" * 78)
    pr("D. STRICT-SAFE vs NC-CONDITIONAL SPLIT")
    pr("=" * 78)
    strict = [r for r in yes_rows if r["license"] in STRICT_SAFE]
    nc = [r for r in yes_rows if r["license"] in NC_VARIANTS]
    other = [r for r in yes_rows if r["license"] not in STRICT_SAFE | NC_VARIANTS]
    pr(f"Strict-safe (CC-BY, CC-BY-SA, CC0, PD): {len(strict)}")
    pr(f"NC-conditional (CC-BY-NC, CC-BY-NC-SA): {len(nc)}")
    pr(f"Other / unrecognized:                   {len(other)}")
    pr("")
    pr("INTERPRETATION:")
    pr("  - Strict-safe papers are unconditionally OK to ship in any package")
    pr("    (MIT, Apache, GPL, anything). Maximum legal safety.")
    pr("  - NC-conditional papers are only OK if your package is licensed")
    pr("    such that downstream users CANNOT use it commercially. If you")
    pr("    license the package as MIT/Apache (which permit commercial use),")
    pr("    a strict reading makes bundling NC content a license conflict.")
    pr("  - 'Other' papers have unusual licenses; review each by hand.")
    pr("")

    if other:
        pr("Other-license papers (review individually):")
        for r in other:
            pr(f"  [{r['license']!r}]  {safe(r['filename'])[:90]}")
        pr("")

    # ----- Check E: Per-license inventory -----
    pr("=" * 78)
    pr("E. PER-LICENSE INVENTORY")
    pr("=" * 78)
    by_lic = defaultdict(list)
    for r in yes_rows:
        by_lic[r["license"]].append(r)
    for lic in sorted(by_lic, key=lambda k: (-len(by_lic[k]), k)):
        rs = by_lic[lic]
        pr(f"\n  {lic} ({len(rs)} papers)")
        for r in sorted(rs, key=lambda r: r["filename"].lower())[:5]:
            pr(f"    {safe(r['filename'])[:90]}")
        if len(rs) > 5:
            pr(f"    ... and {len(rs) - 5} more")
    pr("")

    # ----- Recommendation -----
    pr("=" * 78)
    pr("RECOMMENDATION FOR PUBLISHING (conservative reading)")
    pr("=" * 78)
    pr(f"Maximum legal safety  -> ship only strict-safe ({len(strict)} papers)")
    pr(f"  Tradeoff: lose {len(nc) + len(other)} NC-conditional papers")
    pr(f"  These are CC-licensed but with NC clause that may conflict with")
    pr(f"  a commercially-friendly package license like MIT/Apache.")
    pr("")
    pr(f"Strategy B as configured -> ship {len(yes_rows)} (strict + NC variants)")
    pr(f"  Tradeoff: requires the package license to be NC-compatible, OR")
    pr(f"  a NOTICE file documenting that bundled papers carry NC restrictions")
    pr(f"  that downstream users must respect independently of the code license.")
    pr("")

    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {OUT_TXT}")

    # Build the strict-safe xlsx
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in ["Journal Articles", "Book Chapters"]:
        ws = wb.create_sheet(sheet_name)
        # Header
        cols = [("license", 14), ("filename", 60), ("title", 60),
                ("year", 8), ("doi", 32), ("oa_status", 10), ("category", 30)]
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        for ci, (n, w) in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=n)
            c.fill = header_fill
            c.font = Font(bold=True)
            ws.column_dimensions[c.column_letter].width = w
        # Data
        wrap = Alignment(wrap_text=True, vertical="top")
        sheet_rows = sorted(
            (r for r in strict if r["sheet"] == sheet_name),
            key=lambda r: r["filename"].lower(),
        )
        for ri, r in enumerate(sheet_rows, 2):
            for ci, (n, _) in enumerate(cols, 1):
                v = r.get(n)
                if isinstance(v, str):
                    v = xml_safe(v)
                ws.cell(row=ri, column=ci, value=v).alignment = wrap
        ws.freeze_panes = "A2"
    wb.save(OUT_STRICT_XLSX)
    print(f"Wrote {OUT_STRICT_XLSX} (strict-safe only, {len(strict)} papers)")


if __name__ == "__main__":
    main()
