"""Integrate user's manually-extracted keywords into the .bib file.

Reads the user's Excel + my filename->bib_key CSV, then rewrites the .bib:
  - REPLACE keywords for 241 papers where user extracted them (with
    cleanups: strip trailing periods, split Rulli 2019's
    'Water scarcity, CO2 emissions' into two items).
  - DELETE keywords for 16 papers whose existing .bib keywords are
    either MeSH auto-indexing junk (8) or placeholder `telecoupling`-
    only (8).
  - PRESERVE keywords for 3 legit-looking entries and everything else.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl

BIB_PATH = Path("src/metacoupling/data/telecoupling_literature.bib")
XLSX_PATH = Path("D:/Check/journal_research_keywords_extracted.xlsx")
CSV_PATH = Path("tmp/keyword_extraction.csv")

# Entries to delete the keywords field from (16 total)
DELETE_KEYS = {
    # MeSH-junk (8): contain article/human/animal/controlled-study/etc.
    "booth_investigating_2021",
    "chen_physical_2023",
    "dasilva_socioeconomic_2021",
    "fastner_telecoupled_2023",
    "green_linking_2019",
    "gurney_redefining_2017",
    "sun_importing_2018",
    "xu_impacts_2020",
    # Placeholder-only (8): just `telecoupling`
    "barbieri_food_2022",
    "chai_telecoupled_2024",
    "henriksson_measuring_2018",
    "jiren_multi_2022",
    "scott_remittances_2024",
    "strecker_kenya_2023",
    "zhang_remittance_2024",
    "zhao_tightening_2019",
}

# Per-paper cleanups applied to user's extracted keyword list before insertion.
# Key = bib_key, value = (find_str, replace_str) applied on the semicolon-
# joined raw string.
SPECIFIC_CLEANUPS = {
    # Rulli 2019: split `Water scarcity, CO2 emissions` into two separate items
    "rulli_interdependencies_2019": ("Water scarcity, CO2 emissions",
                                     "Water scarcity; CO2 emissions"),
}


def load_user_keywords() -> dict[str, str]:
    """Return {filename: raw_keyword_string} from the user's Excel."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    sheet = wb["All Journal Research"]
    header = [c.value for c in next(sheet.iter_rows())]
    out = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        if d.get("filename") and d.get("keywords"):
            out[d["filename"]] = d["keywords"]
    return out


def load_filename_to_bibkey() -> dict[str, str]:
    out = {}
    with open(CSV_PATH, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["bib_key"]:
                out[r["filename"]] = r["bib_key"]
    return out


def clean_user_keywords(raw: str, bib_key: str) -> str:
    """Apply cleanups: strip trailing periods on last keyword,
    per-paper fixups, etc."""
    # Apply per-paper specific cleanups first
    if bib_key in SPECIFIC_CLEANUPS:
        find, replace = SPECIFIC_CLEANUPS[bib_key]
        raw = raw.replace(find, replace)

    # Split on ;
    parts = [p.strip() for p in raw.split(";") if p.strip()]
    # Strip trailing periods from each item
    parts = [p.rstrip(".").strip() for p in parts]
    # Drop empties again after stripping
    parts = [p for p in parts if p]
    return "; ".join(parts)


def rewrite_entry(entry_text: str, key: str, new_kw: str | None,
                  action: str) -> tuple[str, bool]:
    """Return (new_entry_text, changed_flag).

    - action='replace' with new_kw != None: replace or insert keywords.
    - action='delete': remove the keywords field entirely.
    - action='preserve' or other: return unchanged.
    """
    if action == "preserve":
        return entry_text, False

    kw_pattern = re.compile(
        r"(?P<indent>\n\s+)keywords\s*=\s*\{[^}]*\},?",
        re.IGNORECASE,
    )
    m = kw_pattern.search(entry_text)

    if action == "delete":
        if not m:
            return entry_text, False
        # Remove the whole line including its leading newline+indent
        new_text = entry_text[:m.start()] + entry_text[m.end():]
        return new_text, True

    if action == "replace":
        assert new_kw is not None
        new_field = f"keywords = {{{new_kw}}},"
        if m:
            # Replace existing keywords line, preserve the indent
            indent = m.group("indent")
            replacement = f"{indent}{new_field}"
            new_text = entry_text[:m.start()] + replacement + entry_text[m.end():]
            return new_text, True
        else:
            # Insert before the last `}` (closing brace) of the entry.
            # Find the final `}` preceded by newline.
            close_match = re.search(r"(\n\s*\})\s*$", entry_text)
            if not close_match:
                return entry_text, False  # can't find closing brace safely
            # Determine a sensible indent from surrounding fields
            prev_field = re.search(r"\n(\s+)\w+\s*=", entry_text)
            indent = prev_field.group(1) if prev_field else "  "
            # Ensure the previous non-brace line ends with ','
            before = entry_text[:close_match.start()]
            before_stripped = before.rstrip()
            if not before_stripped.endswith(","):
                before = before_stripped + ","
            after = entry_text[close_match.start():]
            new_text = before + f"\n{indent}{new_field}" + after
            return new_text, True

    return entry_text, False


def main():
    user_kw = load_user_keywords()
    fname_to_bibkey = load_filename_to_bibkey()

    # Build bib_key -> user_kw_string
    bibkey_to_user_kw: dict[str, str] = {}
    for fname, raw in user_kw.items():
        bibkey = fname_to_bibkey.get(fname)
        if bibkey:
            bibkey_to_user_kw[bibkey] = raw

    print(f"User keywords mapped to .bib keys: {len(bibkey_to_user_kw)}")
    print(f"Keys flagged for DELETE: {len(DELETE_KEYS)}")
    print()

    with open(BIB_PATH, encoding="utf-8", newline="\n") as f:
        bib_text = f.read()

    # Parse into entries, preserving between-entry whitespace
    starts = [m.start() for m in re.finditer(r"^@\w+\{", bib_text, re.MULTILINE)]
    out_parts: list[str] = []
    cursor = 0
    stats = {"replace": 0, "delete": 0, "preserve": 0, "not_found": 0}

    for i, start in enumerate(starts):
        end = starts[i + 1] if i + 1 < len(starts) else len(bib_text)
        # Trim trailing whitespace that would belong to next entry's separator
        chunk = bib_text[start:end]
        key_m = re.match(r"@\w+\{\s*([^,]+)\s*,", chunk)
        out_parts.append(bib_text[cursor:start])
        if not key_m:
            out_parts.append(chunk)
            cursor = end
            continue
        key = key_m.group(1).strip()

        # Find trailing whitespace of this entry (blank lines between entries)
        # The entry itself ends at its closing `}`; trailing content is
        # whitespace until the next `@`.
        entry_end_m = re.search(r"\n\}\s*(\n|\Z)", chunk)
        if entry_end_m:
            entry_text = chunk[:entry_end_m.start() + 2]  # include '\n}'
            trailing = chunk[entry_end_m.start() + 2:]
        else:
            entry_text = chunk
            trailing = ""

        # Decide action
        if key in bibkey_to_user_kw:
            cleaned = clean_user_keywords(bibkey_to_user_kw[key], key)
            new_entry, changed = rewrite_entry(entry_text, key, cleaned, "replace")
            stats["replace" if changed else "preserve"] += 1
        elif key in DELETE_KEYS:
            new_entry, changed = rewrite_entry(entry_text, key, None, "delete")
            stats["delete" if changed else "preserve"] += 1
        else:
            new_entry, _ = rewrite_entry(entry_text, key, None, "preserve")
            stats["preserve"] += 1

        out_parts.append(new_entry + trailing)
        cursor = end

    out_parts.append(bib_text[cursor:])
    new_bib = "".join(out_parts)

    with open(BIB_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write(new_bib)

    print(f"REPLACE (keywords updated): {stats['replace']}")
    print(f"DELETE  (keywords removed): {stats['delete']}")
    print(f"PRESERVE (unchanged):       {stats['preserve']}")
    print(f"NOT_FOUND (no @article):    {stats['not_found']}")
    print()
    print(f"Wrote {BIB_PATH}")


if __name__ == "__main__":
    main()
